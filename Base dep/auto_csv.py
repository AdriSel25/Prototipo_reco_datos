import openpyxl
import funciones as f
import csv
from os import path

print("Ingrese el nombre o direaccion del Archivo Excel a Evaluar (Ej. 'Excel DEP', '../Archivo'):")
nombre_archivo = input('')
nombre_excel = nombre_archivo + '.xlsx'

if not(path.exists(nombre_archivo)):
    print("Ese archivo no existe!")
    exit(1)

#Excel
wb = openpyxl.load_workbook(nombre_excel, data_only=True)
first_sheet = wb[wb.sheetnames[0]]
second_sheet = wb[wb.sheetnames[1]]
wb.active = 1
print(wb.active)

#Listas
datos_depuracion = []
datos_incompletos = []
admin_script = []
nombres_comercios = []
relaciones_procesadas = []
diccionario_admins = {}

#Diccionario
for row in range(2, second_sheet.max_row + 1):
    admin_row = "{}{}".format("A", row)
    comercios_row = "{}{}".format("B", row)

    valor_admin = second_sheet[admin_row].value
    valor_comercios = second_sheet[comercios_row].value

    if (fd.regex_UUID(valor_admin) and fd.regex_UUID(valor_comercios)):
        if valor_admin in diccionario_admins :
            print(diccionario_admins[valor_admin])
            diccionario_admins[valor_admin].append(valor_comercios)
        else:
            diccionario_admins[valor_admin] = [valor_comercios]
    else:
        print("DICCIONARIO: parece que no podemos leer los datos.")

print(diccionario_admins)

def obtener_comercios(fila):
    celda_c = "{}{}".format("P", fila)
    id_comercio = first_sheet[celda_c].value
    return id_comercio

def obtener_admin(fila):
    celda_a = "{}{}".format("Q", fila)
    id_admin = first_sheet[celda_a].value
    return id_admin

#For en donde se recorren la sfila y se leen las marcas en el documento excel

wb.active = 0
print(wb.active)

for row in range(2, first_sheet.max_row +1):

    nombe_comercio_cell = "{}{}".format("A",row)
    comercio_x = "{}{}".format("L",row)
    admin_x = "{}{}".format("M",row)
    relacion_x = "{}{}".format("N",row)
    cantidad_comercios_admin = "{}{}".format("I",row)

    nombre_comercio = str(first_sheet[nombe_comercio_cell].value)
    valor_comercio_x = str(first_sheet[comercio_x].value)
    valor_admin_x = str(first_sheet[admin_x].value)
    valor_relacion_x = str(first_sheet[relacion_x].value)
    valor_cantidad_CA = str(first_sheet[cantidad_comercios_admin].value)

    admin_id = obtener_admin(row)
    comercio_id = obtener_comercios(row)


    if (valor_comercio_x == "X" and valor_admin_x is None and fd.regex_UUID(comercio_id)):
        print(f"Datos Dila {row} \nHay que borrar solo el comercio! Obtendremos el UUID del comercio\n")

        datos_depuracion.append([comercio_id])
    
    elif (valor_comercio_x is None and valor_admin_x is None and valor_relacion_x == "X" 
          and (fd.regex_UUID(comercio_id) and fd.regex_UUID(admin_id))):
        
        print(f"Datos Dila {row} \nHay que borrar solo la relacion! Obtendremos el UUID del comercio y de admin\n")

        datos_depuracion.append([comercio_id, admin_id])

    elif (valor_comercio_x == "X" and valor_admin_x == "X" and valor_cantidad_CA == "NO"
          and fd.regex_UUID(comercio_id)):
        
        print(f"Datos Dila {row} \nHay que borrar el admin y el comercio! Obtendremos el UUID del comercio\n")

        datos_depuracion.append([comercio_id])

    elif (valor_comercio_x == "X" and valor_admin_x == "X" and valor_cantidad_CA == "SI" 
          and (fd.regex_UUID(comercio_id) and fd.regex_UUID(admin_id))):
        
        print(f"Datos Dila {row} \nVoy a buscar en el diccionario_admins...")

        if (admin_id in diccionario_admins and comercio_id in diccionario_admins[admin_id]):
            print("Tengo el admin y el comercio\n")
            fd.procesar(admin_id, diccionario_admins, relaciones_procesadas, datos_depuracion)
            datos_depuracion.append([comercio_id])
            print(f"Se borraraa el comercio: {comercio_id}\n")

        elif valor_admin in diccionario_admins :
            print("tengo el admin pero no el comercio, Por favor revisar excel\n")
            datos_incompletos.append([comercio_id])
            print(f"Se borrara el comercio: {comercio_id}\n")

        else:
            print("No tengo ni el admin ni al comercio, Por favor revisar el excel\n")
            datos_incompletos.append([row, nombre_comercio, comercio_id, admin_id])

    elif (valor_comercio_x is None and valor_admin_x == "X" and fd.regex_UUID(admin_id)):
        if (valor_cantidad_CA == "SI"):
            print(f"Datos Dila {row} \nVoy a buscar en el diccionario_admins para borrar el admin...")

            if (admin_id in diccionario_admins):
                print("Tengo el admin\n")
                fd.procesar(admin_id, diccionario_admins, relaciones_procesadas, datos_depuracion)
            else:
                print("Este admn no esta en el diccionario_admin, por favor revisar Excel\n")
                datos_incompletos.append([row, nombre_comercio, comercio_id, admin_id])

        else:
            print(f"Datos Fila {row} \nHay que borrar el admin! solo borrareos la relacion con el uuid de comercio y admin...")
            print(f"Admin: {admin_id}. Longitud: {len(admin_id)}")
            datos_depuracion.append([comercio_id, admin_id])

    else:
        print(f"Datos Dila {row} \nMe parece que me faltan indicacciones...Puede que las celdas esten vacios o los datos no cumplan con el formato requerido\n")
        datos_incompletos.append([row, nombre_comercio, comercio_id, admin_id])

fd.creacion_csv(datos_depuracion, nombre_archivo, datos_incompletos)
