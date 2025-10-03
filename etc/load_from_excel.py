import openpyxl
from openpyxl.utils import get_column_letter
from os import path
from datetime import datetime
import uuid

from repo import (
    crear_personal_client,
    crear_commercial_client,
    ValidationError
)

# Mapea columnas de la hoja origen (ajústalas si cambia tu layout)
COL_UUID   = 'A'
COL_TYPE   = 'B'  # 'personal' | 'comercial' | 'emprendedor'
COL_PID    = 'C'  # personal_identification
COL_FNAME  = 'D'
COL_LNAME  = 'E'
COL_EMAIL  = 'F'
COL_PHONE  = 'G'
COL_CNAME  = 'H'  # commerce_name
COL_RUC    = 'I'
COL_ADMIN  = 'J'
COL_REGDT  = 'K'  # registration_date

# Nombres de hojas destino (clasificadas)
SHEET_PERSONAL   = 'personal_client'
SHEET_COMMERCIAL = 'commercial_client'

def read_cell(sheet, col, row):
    return sheet[f'{col}{row}'].value

def write_cell(sheet, col, row, value):
    sheet[f'{col}{row}'].value = value

def coerce_excel_datetime(v):
    if v is None:
        return datetime.now()
    if isinstance(v, datetime):
        return v
    return str(v)  # el repo intentará parsear strings

def ensure_replace_sheet(wb, title):
    """Crea la hoja title; si existe, la elimina y la recrea."""
    if title in wb.sheetnames:
        ws = wb[title]
        wb.remove(ws)
    return wb.create_sheet(title)

def autosize_columns(ws):
    widths = {}
    for row in ws.iter_rows(values_only=True):
        for i, val in enumerate(row, start=1):
            length = len(str(val)) if val is not None else 0
            widths[i] = max(widths.get(i, 0), length)
    for i, w in widths.items():
        ws.column_dimensions[get_column_letter(i)].width = min(max(w + 2, 10), 50)

def main():
    print("Ingrese el nombre o dirección del Archivo Excel (sin extensión):")
    nombre_archivo = input('').strip()
    src_path = nombre_archivo + '.xlsx'
    out_path = nombre_archivo + '_uuid.xlsx'  # copia de salida con UUID + hojas clasificadas

    if not path.exists(src_path):
        print("Ese archivo no existe.")
        return

    wb = openpyxl.load_workbook(src_path, data_only=True)
    sheet = wb[wb.sheetnames[0]]

    print("Leyendo registros del Excel e insertando en BD...\n")

    ok, fail, generated = 0, 0, 0

    # Buffers para escribir hojas clasificadas
    personal_rows = []   # (client_uuid, personal_identification, first_name, last_name, email, phone_number, registration_date)
    commercial_rows = [] # (client_uuid, commerce_name, ruc, admin_id, email, registration_date)

    for row in range(2, sheet.max_row + 1):
        client_type = (str(read_cell(sheet, COL_TYPE, row) or '')).strip().lower()

        # UUID: si falta, generar y escribir de vuelta a la hoja origen
        raw_uuid = str(read_cell(sheet, COL_UUID, row) or '').strip()
        client_uuid = raw_uuid if raw_uuid else str(uuid.uuid4())
        if not raw_uuid:
            write_cell(sheet, COL_UUID, row, client_uuid)
            generated += 1

        try:
            if client_type == 'personal':
                personal_identification = str(read_cell(sheet, COL_PID,   row) or '').strip()
                first_name              = str(read_cell(sheet, COL_FNAME, row) or '').strip()
                last_name               = str(read_cell(sheet, COL_LNAME, row) or '').strip()
                email                   = str(read_cell(sheet, COL_EMAIL, row) or '').strip()
                phone                   = str(read_cell(sheet, COL_PHONE, row) or '').strip()
                reg_date                = coerce_excel_datetime(read_cell(sheet, COL_REGDT, row))

                crear_personal_client(
                    client_uuid=client_uuid,
                    personal_identification=personal_identification,
                    first_name=first_name,
                    last_name=last_name,
                    email=email,
                    phone_number=phone,
                    registration_date=reg_date
                )
                # Guardar fila para hoja clasificada
                personal_rows.append((
                    client_uuid, personal_identification, first_name, last_name,
                    email, phone, reg_date
                ))
                print(f"✔ [PERSONAL] Fila {row} insertada (UUID={client_uuid})")
                ok += 1

            elif client_type in ('comercial', 'emprendedor', 'emprendimiento'):
                commerce_name = str(read_cell(sheet, COL_CNAME, row) or '').strip()
                ruc           = str(read_cell(sheet, COL_RUC,   row) or '').strip()
                admin_id      = str(read_cell(sheet, COL_ADMIN, row) or '').strip()
                email         = str(read_cell(sheet, COL_EMAIL, row) or '').strip()
                reg_date      = coerce_excel_datetime(read_cell(sheet, COL_REGDT, row))

                crear_commercial_client(
                    client_uuid=client_uuid,
                    commerce_name=commerce_name,
                    ruc=ruc,
                    admin_id=admin_id,
                    email=email,
                    registration_date=reg_date
                )
                # Guardar fila para hoja clasificada
                commercial_rows.append((
                    client_uuid, commerce_name, ruc, admin_id, email, reg_date
                ))
                print(f"✔ [COMERCIAL] Fila {row} insertada (UUID={client_uuid})")
                ok += 1

            else:
                print(f"↷ [DESCONOCIDO] Fila {row}: tipo '{client_type}' no reconocido. Saltado.")
                fail += 1

        except ValidationError as ve:
            print(f"✖ [Fila {row}] Validación: {ve}")
            fail += 1
        except Exception as e:
            print(f"✖ [Fila {row}] Error inesperado: {e}")
            fail += 1

    # ─────────────────────────────────────────────────────────
    # Crear/rehacer hojas clasificadas y escribir datos
    # ─────────────────────────────────────────────────────────
    ws_personal = ensure_replace_sheet(wb, SHEET_PERSONAL)
    ws_commercial = ensure_replace_sheet(wb, SHEET_COMMERCIAL)

    # Encabezados
    ws_personal.append(("client_uuid","personal_identification","first_name","last_name","email","phone_number","registration_date"))
    ws_commercial.append(("client_uuid","commerce_name","ruc","admin_id","email","registration_date"))

    # Escribir filas
    for r in personal_rows:
        # formateo de fecha amable en Excel
        row = list(r)
        if isinstance(row[-1], datetime):
            row[-1] = row[-1].strftime("%Y-%m-%d %H:%M:%S")
        ws_personal.append(tuple(row))

    for r in commercial_rows:
        row = list(r)
        if isinstance(row[-1], datetime):
            row[-1] = row[-1].strftime("%Y-%m-%d %H:%M:%S")
        ws_commercial.append(tuple(row))

    # Autoajustar columnas
    autosize_columns(ws_personal)
    autosize_columns(ws_commercial)

    # Guardar copia con UUID + hojas clasificadas
    wb.save(out_path)

    print(f"\nTerminado.")
    print(f"  • Insertados OK : {ok}")
    print(f"  • Fallidos/omit.: {fail}")
    print(f"  • UUID generados: {generated}")
    print(f"  • Archivo de salida (con hojas clasificadas): {out_path}")

if __name__ == "__main__":
    main()
